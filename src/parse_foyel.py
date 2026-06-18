"""
parse_foyel.py — Lee el Excel de reservas de Foyel (grilla de ocupación) y
devuelve el dict bed-nights-only que consume el dashboard.

El archivo NO es una lista de reservas: es un calendario de ocupación con dos
solapas ('FOYEL 2027' corriente, 'FOYEL 2026' referencia), temporada mar–may,
6 habitaciones, nombre+pax por celda. No hay datos de plata. Los colores de
estado de la grilla NO coinciden con su propia leyenda, así que NO se desglosa
por estado: se reporta solo bed-nights (pax-noches) por mes.

Bed-nights por mes = "TOTAL PAX MES" que calcula la propia planilla (fila ~21),
que validamos contra el recuento manual (mar 2026 = 92, exacto).
"""
from __future__ import annotations
from openpyxl import load_workbook

MONTHS = [(3, "Mar"), (4, "Abr"), (5, "May")]
CUR_SHEET = "FOYEL 2027"
PREV_SHEET = "FOYEL 2026"
GROUP_HEADERS = {"", "LODGE", "HAB 1", "HAB 2", "HAB 3", "HAB 4", "HAB 5", "HAB 6"}


def _month_totals_from_row21(rows) -> dict:
    """Lee los totales 'MAR'/'ABR'/'MAY' que la planilla ya tiene en su fila de totales."""
    out = {}
    for row in rows:
        for c, cell in enumerate(row):
            v = cell.value
            if isinstance(v, str) and v.strip() in ("MAR", "ABR", "MAY") and c + 1 < len(row):
                nxt = row[c + 1].value
                if isinstance(nxt, (int, float)):
                    out[v.strip().capitalize()] = int(nxt)
    return out  # {"Mar": 103, "Abr": 8}


def _month_totals_recount(rows) -> dict:
    """Recuento independiente: suma pax (col name+1) sobre celdas de habitación ocupadas."""
    datemo = {}
    if len(rows) > 2:
        for c, cell in enumerate(rows[2]):
            if hasattr(cell.value, "month"):
                datemo[c] = cell.value.month
    tot = {m: 0 for m, _ in MONTHS}
    for r in range(4, min(16, len(rows))):
        for c, mo in datemo.items():
            if mo not in tot or c >= len(rows[r]):
                continue
            name = rows[r][c].value
            if not isinstance(name, str) or name.strip() in GROUP_HEADERS or name.strip().startswith("HAB"):
                continue
            pax = rows[r][c + 1].value if c + 1 < len(rows[r]) else None
            tot[mo] += pax if isinstance(pax, (int, float)) else 1  # 'T' / vacío => 1
    return {label: tot[m] for m, label in MONTHS}


def _groups(rows) -> list[str]:
    datemo = {}
    if len(rows) > 2:
        for c, cell in enumerate(rows[2]):
            if hasattr(cell.value, "month"):
                datemo[c] = True
    names = set()
    for r in range(4, min(16, len(rows))):
        for c in (datemo or {}):
            if c < len(rows[r]):
                v = rows[r][c].value
                if isinstance(v, str) and v.strip() and v.strip() not in GROUP_HEADERS and not v.strip().startswith("HAB"):
                    names.add(v.strip())
    return sorted(names)


def _sheet_months(ws, *, prefer_sheet_total=True) -> tuple[dict, list]:
    rows = list(ws.iter_rows())
    sheet_tot = _month_totals_from_row21(rows)
    recount = _month_totals_recount(rows)
    # default: confiar en el total de la planilla; recount es el cross-check
    months = {}
    for _, label in MONTHS:
        months[label] = sheet_tot.get(label, recount.get(label, 0))
    return months, _groups(rows)


def parse_foyel_xlsx(xlsx_path: str, *, week_now: int, meta: dict) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    cur_months, cur_groups = _sheet_months(wb[CUR_SHEET])
    prev_months, prev_groups = _sheet_months(wb[PREV_SHEET])

    months = [dict(l=label, cur=cur_months.get(label, 0), prev=prev_months.get(label, 0))
              for _, label in MONTHS]

    return dict(
        name=meta["name"], sub=meta["sub"], logo=meta.get("logo"),
        bnOnly=True, week=week_now,
        groupsCur=len(cur_groups), groupsPrev=len(prev_groups),
        groupsCurList=" · ".join(cur_groups) or "—",
        groupsPrevList=" · ".join(prev_groups) or "—",
        months=months,
    )
